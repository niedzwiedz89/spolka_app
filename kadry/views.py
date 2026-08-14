from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.contrib import messages as django_messages
from django.contrib.contenttypes.models import ContentType
from django.http import HttpResponse
import io
import calendar
import xml.etree.ElementTree as ET
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import (
    Pracownik, Budowa, PracownikBudowa, Mieszkanie, Pojazd, ZdarzenieFloty,
    Zalacznik, Faktura, FakturaXml,
)
from .forms import (
    PracownikForm, BudowaForm, PracownikBudowaForm,
    MieszkanieForm, PojazdForm, ZdarzenieFlotyForm, ZalacznikForm
)
from .ksef_xml import pozycje_faktury, wczytaj_xml


# =============================================================================
# Helper function for wildcard filtering
# =============================================================================

def get_wildcard_filter(field_name, value):
    """
    Zwraca obiekt Q na podstawie dopasowań ze znakami %:
    - '%' – po prostu zignoruj puste
    - 'M%ch%' – regex lub złożone dopasowania (tutaj dla uproszczenia rzutowane na contains)
    - 'Michal%' -> startswith
    - '%chal' -> endswith
    - '%Michal%' -> icontains
    - 'Michal' -> exact/icontains jako domyślne zachowanie (zachowujemy icontains dla kompatybilności wstecz)
    """
    if not value or value == '%':
        return Q()

    if value.startswith('%') and value.endswith('%'):
        # %tekst%
        stripped = value.strip('%')
        return Q(**{f"{field_name}__icontains": stripped})
    elif value.startswith('%'):
        # %tekst
        stripped = value.lstrip('%')
        return Q(**{f"{field_name}__iendswith": stripped})
    elif value.endswith('%'):
        # tekst%
        stripped = value.rstrip('%')
        return Q(**{f"{field_name}__istartswith": stripped})
    elif '%' in value:
        # np. M%ch%l - regex dla SQLite
        pattern = value.replace('%', '.*')
        return Q(**{f"{field_name}__iregex": f"^{pattern}$"})
    else:
        # domyślnie, bez %
        return Q(**{f"{field_name}__icontains": value})


# =============================================================================
# DASHBOARD
# =============================================================================

def dashboard(request):
    """Strona główna – pulpit z kafelkami modułów."""
    stats = {
        "pracownicy": Pracownik.objects.count(),
        "budowy": Budowa.objects.count(),
        "mieszkania": Mieszkanie.objects.count(),
        "pojazdy": Pojazd.objects.count(),
        "faktury": Faktura.objects.count(),
    }
    return render(request, "kadry/dashboard.html", {"stats": stats})


# =============================================================================
# PRACOWNICY
# =============================================================================

def pracownik_list(request):
    """Lista pracowników z filtrami."""
    qs = Pracownik.objects.all()

    # Filtry
    f_imie = request.GET.get("imie", "")
    f_nazwisko = request.GET.get("nazwisko", "")
    f_pesel = request.GET.get("pesel", "")
    f_telefon = request.GET.get("telefon", "")

    if f_imie:
        qs = qs.filter(get_wildcard_filter("imie", f_imie))
    if f_nazwisko:
        qs = qs.filter(get_wildcard_filter("nazwisko", f_nazwisko))
    if f_pesel:
        qs = qs.filter(get_wildcard_filter("pesel", f_pesel))
    if f_telefon:
        qs = qs.filter(get_wildcard_filter("telefon", f_telefon))

    paginator = Paginator(qs, 50)
    page = request.GET.get("strona", 1)
    objects = paginator.get_page(page)

    # ------------------------------------------------------------------
    # Obliczanie alertów wierszy — podświetlenie pracowników wymagających uwagi
    # Poziomy: "danger" (wygasło / brak A1), "warning" (kończy się ≤ 30 dni)
    # ------------------------------------------------------------------
    dzisiaj = date.today()
    prog_ostrzezenia = 30  # dni przed wygaśnięciem A1

    row_alerts = {}  # {pk: "danger"|"warning"}
    for p in objects:
        poziom = None
        if p.a1_do is not None:
            dni_do_konca = (p.a1_do - dzisiaj).days
            if dni_do_konca < 0:
                # A1 już wygasła
                poziom = "danger"
            elif dni_do_konca <= prog_ostrzezenia:
                # A1 kończy się za mniej niż 30 dni
                poziom = "warning"
        elif p.a1_od is not None and p.a1_do is None:
            # Ustawiono datę od, ale brak daty do — traktuj jako brak ważnego dokumentu
            poziom = "warning"
        if poziom:
            row_alerts[p.pk] = poziom

    context = {
        "objects": objects,
        "module_name": "Pracownicy",
        "row_alerts": row_alerts,
        "filters": {
            "imie": f_imie,
            "nazwisko": f_nazwisko,
            "pesel": f_pesel,
            "telefon": f_telefon,
        },
        "columns": [
            ("imie", "Imię"),
            ("nazwisko", "Nazwisko"),
            ("pesel", "PESEL"),
            ("get_status_display", "Status"),
            ("telefon", "Telefon"),
            ("email", "E-mail"),
            ("dowod_osobisty", "Dowód osobisty"),
        ],
        "detail_url_name": "pracownik_detail",
        "create_url_name": "pracownik_create",
        "delete_url_name": "pracownik_delete",
        "bulk_actions": [
            {
                "id": "raport_obecnosci",
                "label": "📊 Generuj raport obecności (Excel)",
                "url": reverse("pracownik_raport_budowy"),
                "id_field": "pracownik_ids",
                "needs_month": True,
            }
        ],
    }
    return render(request, "kadry/list.html", context)


def pracownik_raport_budowy(request):
    """Widok obsługujący POST z masowej akcji generowania raportu budów."""
    if request.method == "POST":
        pracownik_ids = request.POST.getlist("pracownik_ids")
        miesiac_str = request.POST.get("raport_miesiac") # Np. 2026-02

        if not pracownik_ids:
            django_messages.error(request, "Nie wybrano żadnych pracowników.")
            return redirect("pracownik_list")

        # Ustalenie roku i miesiąca
        try:
            year, month = map(int, miesiac_str.split('-'))
        except (ValueError, AttributeError):
            dzisiaj = date.today()
            year, month = dzisiaj.year, dzisiaj.month

        _, num_days = calendar.monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, num_days)

        pracownicy = list(Pracownik.objects.filter(pk__in=pracownik_ids).order_by("nazwisko", "imie"))
        if not pracownicy:
            django_messages.error(request, "Brak pasujących pracowników w bazie.")
            return redirect("pracownik_list")

        # Pobieramy przypisania z wybranego miesiąca
        # Przypisanie musi trwać (data_od <= end_date) oraz nie może się kończyć przed początkiem (data_do >= start_date albo null)
        przypisania = PracownikBudowa.objects.filter(
            pracownik__in=pracownicy,
            data_od__lte=end_date
        ).filter(
            Q(data_do__isnull=True) | Q(data_do__gte=start_date)
        ).select_related("budowa")

        # Mapujemy dni dla każdego pracownika
        m = {p.id: {} for p in pracownicy}
        for przyp in przypisania:
            p_id = przyp.pracownik.id
            # określamy ramy w obrębie danego miesiąca
            od = max(przyp.data_od, start_date) if przyp.data_od else start_date
            do = min(przyp.data_do, end_date) if przyp.data_do else end_date
            
            for d in range(od.day, do.day + 1):
                # Jeśli jest kilka budów jednego dnia, dopiszemy lub weźmiemy pierwszą
                if d not in m[p_id]:
                    m[p_id][d] = przyp.budowa.nazwa
                else:
                    if przyp.budowa.nazwa not in m[p_id][d]:
                        m[p_id][d] += f", {przyp.budowa.nazwa}"

        # --- Tworzenie pliku Excel ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Raport {month:02d}.{year}"

        # Style wg życzenia użytkownika
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # 1. Nagłówek Pierwszy Wiersz
        ws.cell(row=1, column=1, value="Data\nDatum").font = header_font
        ws.cell(row=1, column=1).alignment = center_align

        # Konfigurujemy pracownikow (2 kolumny na pracownika)
        col_idx = 2
        for p in pracownicy:
            ws.cell(row=1, column=col_idx, value=f"{p.nazwisko} {p.imie}").font = header_font
            ws.cell(row=1, column=col_idx).alignment = center_align
            
            ws.cell(row=1, column=col_idx+1, value="Budowa\nBau-Site").font = header_font
            ws.cell(row=1, column=col_idx+1).alignment = center_align
            
            # Wymiary
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12  # Godziny/Pracownik
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx+1)].width = 16 # Budowa
            
            col_idx += 2

        ws.column_dimensions['A'].width = 12

        # 2. Wiersze z datami w danym miesiącu
        current_row = 2
        for day in range(1, num_days + 1):
            ws.cell(row=current_row, column=1, value=f"{day}.{month:02d}.{year}")
            
            # Dla każdego pracownika odpytujemy mape
            col_idx = 2
            for p in pracownicy:
                # column col_idx to GODZINY - pozostawiamy wolne miejsce by użytkownik wpisał RĘCZNIE potem zero np. D35
                # column col_idx+1 to BUDOWA - wrzucamy Prüm
                nazwa_budowy = m[p.id].get(day, "")
                if nazwa_budowy:
                    ws.cell(row=current_row, column=col_idx+1, value=nazwa_budowy)
                col_idx += 2
                
            current_row += 1

        # 3. Podsumowanie na samym dole (gdzie user trzyma sume z excela)
        # Np 3 rzędy puste jak na screnie po dniach. Dajemy 3 oczka nizej Sumy
        sum_row = current_row + 2
        col_idx = 2
        for p in pracownicy:
            # Wstawienie na twardo 0.00 w polu godzin jak pokazane we wsadzie graficznym uzytkownika row 35
            ws.cell(row=sum_row, column=col_idx, value="0.00").alignment = center_align
            ws.cell(row=sum_row, column=col_idx).fill = gray_fill
            col_idx += 2

        # Zapis i przygotowanie do wysłania
        file_bytes = io.BytesIO()
        wb.save(file_bytes)
        file_bytes.seek(0)

        response = HttpResponse(
            file_bytes.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="RaportBudowy_{month:02d}_{year}.xlsx"'
        return response

    return redirect("pracownik_list")


def pracownik_detail(request, pk):
    """Szczegóły / edycja pracownika."""
    obj = get_object_or_404(Pracownik, pk=pk)
    if request.method == "POST":
        form = PracownikForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            django_messages.success(request, "Zapisano zmiany.")
            return redirect("pracownik_detail", pk=obj.pk)
    else:
        form = PracownikForm(instance=obj)

    # Przypisane budowy
    przypisania = PracownikBudowa.objects.filter(pracownik=obj).select_related("budowa")
    przypisanie_form = PracownikBudowaForm(initial={"pracownik": obj})

    # Załączniki
    ct = ContentType.objects.get_for_model(Pracownik)
    zalaczniki = Zalacznik.objects.filter(content_type=ct, object_id=obj.pk)
    zalacznik_form = ZalacznikForm()

    context = {
        "form": form,
        "object": obj,
        "module_name": "Pracownik",
        "list_url_name": "pracownik_list",
        "przypisania": przypisania,
        "przypisanie_form": przypisanie_form,
        "zalaczniki": zalaczniki,
        "zalacznik_form": zalacznik_form,
        "content_type_id": ct.pk,
    }
    return render(request, "kadry/pracownik_form.html", context)


def pracownik_create(request):
    """Dodawanie nowego pracownika."""
    if request.method == "POST":
        form = PracownikForm(request.POST)
        if form.is_valid():
            obj = form.save()
            django_messages.success(request, "Dodano pracownika.")
            return redirect("pracownik_detail", pk=obj.pk)
    else:
        form = PracownikForm()

    context = {
        "form": form,
        "module_name": "Nowy pracownik",
        "list_url_name": "pracownik_list",
    }
    return render(request, "kadry/pracownik_form.html", context)


def pracownik_delete(request, pk):
    """Usuwanie pracownika."""
    obj = get_object_or_404(Pracownik, pk=pk)
    if request.method == "POST":
        obj.delete()
        django_messages.success(request, "Usunięto pracownika.")
        return redirect("pracownik_list")
    return render(request, "kadry/confirm_delete.html", {
        "object": obj,
        "module_name": "Pracownik",
        "list_url_name": "pracownik_list",
    })


# =============================================================================
# BUDOWY
# =============================================================================

def budowa_list(request):
    """Lista budów z filtrami."""
    qs = Budowa.objects.all()

    f_nazwa = request.GET.get("nazwa", "")
    f_miejsce = request.GET.get("miejsce", "")
    f_adres = request.GET.get("adres", "")

    if f_nazwa:
        qs = qs.filter(get_wildcard_filter("nazwa", f_nazwa))
    if f_miejsce:
        qs = qs.filter(get_wildcard_filter("miejsce", f_miejsce))
    if f_adres:
        qs = qs.filter(get_wildcard_filter("adres", f_adres))

    paginator = Paginator(qs, 50)
    page = request.GET.get("strona", 1)
    objects = paginator.get_page(page)

    context = {
        "objects": objects,
        "module_name": "Budowy",
        "filters": {
            "nazwa": f_nazwa,
            "miejsce": f_miejsce,
            "adres": f_adres,
        },
        "columns": [
            ("nazwa", "Nazwa"),
            ("miejsce", "Miejsce"),
            ("adres", "Adres"),
        ],
        "detail_url_name": "budowa_detail",
        "create_url_name": "budowa_create",
        "delete_url_name": "budowa_delete",
    }
    return render(request, "kadry/list.html", context)


def budowa_detail(request, pk):
    """Szczegóły / edycja budowy."""
    obj = get_object_or_404(Budowa, pk=pk)
    if request.method == "POST":
        form = BudowaForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            django_messages.success(request, "Zapisano zmiany.")
            return redirect("budowa_detail", pk=obj.pk)
    else:
        form = BudowaForm(instance=obj)

    przypisani = PracownikBudowa.objects.filter(budowa=obj).select_related("pracownik")

    context = {
        "form": form,
        "object": obj,
        "module_name": "Budowa",
        "list_url_name": "budowa_list",
        "przypisani": przypisani,
    }
    return render(request, "kadry/budowa_form.html", context)


def budowa_create(request):
    """Dodawanie nowej budowy."""
    if request.method == "POST":
        form = BudowaForm(request.POST)
        if form.is_valid():
            obj = form.save()
            django_messages.success(request, "Dodano budowę.")
            return redirect("budowa_detail", pk=obj.pk)
    else:
        form = BudowaForm()

    context = {
        "form": form,
        "module_name": "Nowa budowa",
        "list_url_name": "budowa_list",
    }
    return render(request, "kadry/budowa_form.html", context)


def budowa_delete(request, pk):
    """Usuwanie budowy."""
    obj = get_object_or_404(Budowa, pk=pk)
    if request.method == "POST":
        obj.delete()
        django_messages.success(request, "Usunięto budowę.")
        return redirect("budowa_list")
    return render(request, "kadry/confirm_delete.html", {
        "object": obj,
        "module_name": "Budowa",
        "list_url_name": "budowa_list",
    })


# =============================================================================
# PRZYPISANIE PRACOWNIK ↔ BUDOWA
# =============================================================================

def przypisanie_create(request, pracownik_pk):
    """Dodawanie przypisania pracownika do budowy."""
    pracownik = get_object_or_404(Pracownik, pk=pracownik_pk)
    if request.method == "POST":
        form = PracownikBudowaForm(request.POST)
        if form.is_valid():
            przypisanie = form.save(commit=False)
            przypisanie.pracownik = pracownik
            przypisanie.save()
            django_messages.success(request, "Przypisano do budowy.")
    return redirect("pracownik_detail", pk=pracownik_pk)


def przypisanie_delete(request, pk):
    """Usuwanie przypisania pracownika do budowy."""
    obj = get_object_or_404(PracownikBudowa, pk=pk)
    pracownik_pk = obj.pracownik.pk
    if request.method == "POST":
        obj.delete()
        django_messages.success(request, "Usunięto przypisanie.")
    return redirect("pracownik_detail", pk=pracownik_pk)


# =============================================================================
# MIESZKANIA
# =============================================================================

def mieszkanie_list(request):
    """Lista mieszkań z filtrami."""
    qs = Mieszkanie.objects.all()

    f_nazwa = request.GET.get("nazwa", "")
    f_miasto = request.GET.get("miasto", "")
    f_adres = request.GET.get("adres", "")

    if f_nazwa:
        qs = qs.filter(get_wildcard_filter("nazwa", f_nazwa))
    if f_miasto:
        qs = qs.filter(get_wildcard_filter("miasto", f_miasto))
    if f_adres:
        qs = qs.filter(get_wildcard_filter("adres", f_adres))

    paginator = Paginator(qs, 50)
    page = request.GET.get("strona", 1)
    objects = paginator.get_page(page)

    context = {
        "objects": objects,
        "module_name": "Mieszkania",
        "filters": {
            "nazwa": f_nazwa,
            "miasto": f_miasto,
            "adres": f_adres,
        },
        "columns": [
            ("nazwa", "Nazwa"),
            ("adres", "Adres"),
            ("miasto", "Miasto"),
            ("pojemnosc", "Pojemność"),
        ],
        "detail_url_name": "mieszkanie_detail",
        "create_url_name": "mieszkanie_create",
        "delete_url_name": "mieszkanie_delete",
    }
    return render(request, "kadry/list.html", context)


def mieszkanie_detail(request, pk):
    """Szczegóły / edycja mieszkania."""
    obj = get_object_or_404(Mieszkanie, pk=pk)
    if request.method == "POST":
        form = MieszkanieForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            django_messages.success(request, "Zapisano zmiany.")
            return redirect("mieszkanie_detail", pk=obj.pk)
    else:
        form = MieszkanieForm(instance=obj)

    context = {
        "form": form,
        "object": obj,
        "module_name": "Mieszkanie",
        "list_url_name": "mieszkanie_list",
    }
    return render(request, "kadry/mieszkanie_form.html", context)


def mieszkanie_create(request):
    """Dodawanie nowego mieszkania."""
    if request.method == "POST":
        form = MieszkanieForm(request.POST)
        if form.is_valid():
            obj = form.save()
            django_messages.success(request, "Dodano mieszkanie.")
            return redirect("mieszkanie_detail", pk=obj.pk)
    else:
        form = MieszkanieForm()

    context = {
        "form": form,
        "module_name": "Nowe mieszkanie",
        "list_url_name": "mieszkanie_list",
    }
    return render(request, "kadry/mieszkanie_form.html", context)


def mieszkanie_delete(request, pk):
    """Usuwanie mieszkania."""
    obj = get_object_or_404(Mieszkanie, pk=pk)
    if request.method == "POST":
        obj.delete()
        django_messages.success(request, "Usunięto mieszkanie.")
        return redirect("mieszkanie_list")
    return render(request, "kadry/confirm_delete.html", {
        "object": obj,
        "module_name": "Mieszkanie",
        "list_url_name": "mieszkanie_list",
    })


# =============================================================================
# FLOTA (POJAZDY)
# =============================================================================

def pojazd_list(request):
    """Lista pojazdów z filtrami."""
    qs = Pojazd.objects.all()

    f_marka = request.GET.get("marka", "")
    f_nr_rej = request.GET.get("nr_rejestracyjny", "")
    f_model = request.GET.get("model_pojazdu", "")

    if f_marka:
        qs = qs.filter(get_wildcard_filter("marka", f_marka))
    if f_nr_rej:
        qs = qs.filter(get_wildcard_filter("nr_rejestracyjny", f_nr_rej))
    if f_model:
        qs = qs.filter(get_wildcard_filter("model_pojazdu", f_model))

    paginator = Paginator(qs, 50)
    page = request.GET.get("strona", 1)
    objects = paginator.get_page(page)

    context = {
        "objects": objects,
        "module_name": "Flota",
        "filters": {
            "marka": f_marka,
            "nr_rejestracyjny": f_nr_rej,
            "model_pojazdu": f_model,
        },
        "columns": [
            ("marka", "Marka"),
            ("model_pojazdu", "Model"),
            ("nr_rejestracyjny", "Nr rejestracyjny"),
            ("przebieg_km", "Przebieg (km)"),
            ("data_przegladu", "Przegląd"),
            ("data_ubezpieczenia", "Ubezpieczenie"),
        ],
        "detail_url_name": "pojazd_detail",
        "create_url_name": "pojazd_create",
        "delete_url_name": "pojazd_delete",
    }
    return render(request, "kadry/list.html", context)


def pojazd_detail(request, pk):
    """Szczegóły / edycja pojazdu + zdarzenia."""
    obj = get_object_or_404(Pojazd, pk=pk)
    if request.method == "POST":
        form = PojazdForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            django_messages.success(request, "Zapisano zmiany.")
            return redirect("pojazd_detail", pk=obj.pk)
    else:
        form = PojazdForm(instance=obj)

    zdarzenia = ZdarzenieFloty.objects.filter(pojazd=obj).order_by("-data")
    zdarzenie_form = ZdarzenieFlotyForm()

    # Załączniki
    ct = ContentType.objects.get_for_model(Pojazd)
    zalaczniki = Zalacznik.objects.filter(content_type=ct, object_id=obj.pk)
    zalacznik_form = ZalacznikForm()

    context = {
        "form": form,
        "object": obj,
        "module_name": "Pojazd",
        "list_url_name": "pojazd_list",
        "zdarzenia": zdarzenia,
        "zdarzenie_form": zdarzenie_form,
        "zalaczniki": zalaczniki,
        "zalacznik_form": zalacznik_form,
        "content_type_id": ct.pk,
    }
    return render(request, "kadry/pojazd_form.html", context)


def pojazd_create(request):
    """Dodawanie nowego pojazdu."""
    if request.method == "POST":
        form = PojazdForm(request.POST)
        if form.is_valid():
            obj = form.save()
            django_messages.success(request, "Dodano pojazd.")
            return redirect("pojazd_detail", pk=obj.pk)
    else:
        form = PojazdForm()

    context = {
        "form": form,
        "module_name": "Nowy pojazd",
        "list_url_name": "pojazd_list",
    }
    return render(request, "kadry/pojazd_form.html", context)


def pojazd_delete(request, pk):
    """Usuwanie pojazdu."""
    obj = get_object_or_404(Pojazd, pk=pk)
    if request.method == "POST":
        obj.delete()
        django_messages.success(request, "Usunięto pojazd.")
        return redirect("pojazd_list")
    return render(request, "kadry/confirm_delete.html", {
        "object": obj,
        "module_name": "Pojazd",
        "list_url_name": "pojazd_list",
    })


# =============================================================================
# ZDARZENIA FLOTY
# =============================================================================

def zdarzenie_create(request, pojazd_pk):
    """Dodawanie zdarzenia do pojazdu."""
    pojazd = get_object_or_404(Pojazd, pk=pojazd_pk)
    if request.method == "POST":
        form = ZdarzenieFlotyForm(request.POST)
        if form.is_valid():
            zdarzenie = form.save(commit=False)
            zdarzenie.pojazd = pojazd
            zdarzenie.save()
            django_messages.success(request, "Dodano zdarzenie.")
    return redirect("pojazd_detail", pk=pojazd_pk)


def zdarzenie_delete(request, pk):
    """Usuwanie zdarzenia."""
    obj = get_object_or_404(ZdarzenieFloty, pk=pk)
    pojazd_pk = obj.pojazd.pk
    if request.method == "POST":
        obj.delete()
        django_messages.success(request, "Usunięto zdarzenie.")
    return redirect("pojazd_detail", pk=pojazd_pk)


# =============================================================================
# ZAŁĄCZNIKI (WIRTUALNE/GENERIC)
# =============================================================================

def _data_lub_none(value):
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return None


def _zakres_dat(miesiac, od, do):
    """Zamienia filtr miesiąca (YYYY-MM) lub parę dat na zakres. Miesiąc ma pierwszeństwo."""
    if miesiac:
        try:
            rok, mies = (int(x) for x in miesiac.split("-"))
            return date(rok, mies, 1), date(rok, mies, calendar.monthrange(rok, mies)[1])
        except (ValueError, TypeError, calendar.IllegalMonthError):
            return None, None
    return _data_lub_none(od), _data_lub_none(do)


def _rola_z_filtra(value):
    """Pozwala filtrować rodzaj faktury po polskiej etykiecie widocznej w tabeli."""
    v = (value or "").strip().strip("%").lower()
    if v and ("zakup".startswith(v) or v.startswith("zakup") or v.startswith("kosz")):
        return "buyer"
    if v and ("sprzedaż".startswith(v) or v.startswith("sprzeda") or v.startswith("sprzedaz")):
        return "seller"
    return None


def faktura_list(request):
    """Lista faktur z KSeF z filtrami."""
    qs = Faktura.objects.all()

    f_numer = request.GET.get("invoice_number", "")
    f_sprzedawca = request.GET.get("seller_name", "")
    f_nip = request.GET.get("seller_nip", "")
    f_typ = request.GET.get("invoice_type", "")
    f_rola = request.GET.get("subject_role", "")
    f_miesiac = request.GET.get("miesiac", "")
    f_od = request.GET.get("data_od", "")
    f_do = request.GET.get("data_do", "")

    data_od, data_do = _zakres_dat(f_miesiac, f_od, f_do)
    if data_od:
        qs = qs.filter(issue_date__gte=data_od)
    if data_do:
        qs = qs.filter(issue_date__lte=data_do)

    if f_numer:
        qs = qs.filter(get_wildcard_filter("invoice_number", f_numer))
    if f_sprzedawca:
        qs = qs.filter(get_wildcard_filter("seller_name", f_sprzedawca))
    if f_nip:
        qs = qs.filter(get_wildcard_filter("seller_nip", f_nip))
    if f_typ:
        qs = qs.filter(get_wildcard_filter("invoice_type", f_typ))
    if f_rola:
        rola = _rola_z_filtra(f_rola)
        qs = qs.filter(subject_role=rola) if rola else qs.filter(
            get_wildcard_filter("subject_role", f_rola)
        )

    paginator = Paginator(qs, 50)
    page = request.GET.get("strona", 1)
    objects = paginator.get_page(page)

    context = {
        "objects": objects,
        "module_name": "Faktury",
        "filters": {
            "invoice_number": f_numer,
            "seller_name": f_sprzedawca,
            "seller_nip": f_nip,
            "invoice_type": f_typ,
            "subject_role": f_rola,
        },
        "columns": [
            ("issue_date", "Data wystawienia"),
            ("invoice_number", "Numer faktury"),
            ("subject_role", "Rodzaj"),
            ("seller_name", "Sprzedawca"),
            ("seller_nip", "NIP sprzedawcy"),
            ("invoice_type", "Typ"),
            ("net_amount", "Netto"),
            ("vat_amount", "VAT"),
            ("gross_amount", "Brutto"),
        ],
        "date_filter": {
            "label": "Data wystawienia",
            "miesiac": f_miesiac,
            "data_od": f_od,
            "data_do": f_do,
        },
        "podsumowanie": qs.aggregate(
            netto=Sum("net_amount"), vat=Sum("vat_amount"), brutto=Sum("gross_amount")
        ),
        "detail_url_name": "faktura_detail",
        "bulk_actions": [
            {
                "id": "eksport_excel",
                "label": "📊 Eksportuj zaznaczone do Excela",
                "url": reverse("faktura_eksport_excel"),
                "id_field": "faktura_ids",
                "needs_month": False,
            }
        ],
    }
    return render(request, "kadry/list.html", context)


def faktura_eksport_excel(request):
    """Eksport zaznaczonych faktur do pliku XLSX."""
    if request.method != "POST":
        return redirect("faktura_list")

    ids = request.POST.getlist("faktura_ids")
    if not ids:
        django_messages.error(request, "Nie wybrano żadnych faktur.")
        return redirect("faktura_list")

    faktury = list(Faktura.objects.filter(pk__in=ids))
    if not faktury:
        django_messages.error(request, "Brak pasujących faktur w bazie.")
        return redirect("faktura_list")

    naglowki = [
        ("Data wystawienia", 16), ("Numer faktury", 26), ("Rodzaj", 10), ("Typ", 8),
        ("Sprzedawca", 42), ("NIP sprzedawcy", 15), ("Nabywca", 42),
        ("Netto", 14), ("VAT", 14), ("Brutto", 14), ("Waluta", 8),
        ("Tryb", 10), ("Numer KSeF", 38),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faktury KSeF"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for kol, (etykieta, szerokosc) in enumerate(naglowki, start=1):
        cell = ws.cell(row=1, column=kol, value=etykieta)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[openpyxl.utils.get_column_letter(kol)].width = szerokosc

    for wiersz, f in enumerate(faktury, start=2):
        ws.cell(row=wiersz, column=1, value=f.issue_date).number_format = "DD.MM.YYYY"
        ws.cell(row=wiersz, column=2, value=f.invoice_number)
        ws.cell(row=wiersz, column=3, value=f.get_subject_role_display())
        ws.cell(row=wiersz, column=4, value=f.invoice_type)
        ws.cell(row=wiersz, column=5, value=f.seller_name)
        ws.cell(row=wiersz, column=6, value=f.seller_nip)
        ws.cell(row=wiersz, column=7, value=f.buyer_name)
        for kol, wartosc in ((8, f.net_amount), (9, f.vat_amount), (10, f.gross_amount)):
            cell = ws.cell(row=wiersz, column=kol,
                           value=float(wartosc) if wartosc is not None else None)
            cell.number_format = "#,##0.00"
        ws.cell(row=wiersz, column=11, value=f.currency)
        ws.cell(row=wiersz, column=12, value=f.invoicing_mode)
        ws.cell(row=wiersz, column=13, value=f.ksef_number)

    suma_wiersz = len(faktury) + 2
    ws.cell(row=suma_wiersz, column=7, value="Razem").font = header_font
    for kol in (8, 9, 10):
        litera = openpyxl.utils.get_column_letter(kol)
        cell = ws.cell(row=suma_wiersz, column=kol,
                       value=f"=SUM({litera}2:{litera}{suma_wiersz - 1})")
        cell.font = header_font
        cell.fill = header_fill
        cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(naglowki))}{suma_wiersz - 1}"

    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)

    response = HttpResponse(
        file_bytes.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="FakturyKSeF_{date.today():%Y%m%d}.xlsx"'
    )
    return response


def faktura_detail(request, pk):
    """Szczegóły faktury z KSeF (tylko do odczytu)."""
    obj = get_object_or_404(Faktura, pk=pk)
    plik_xml = FakturaXml.objects.filter(pk=obj.pk).first()

    pozycje, blad_xml = [], None
    if plik_xml:
        try:
            zawartosc = wczytaj_xml(obj.pk)
            pozycje = pozycje_faktury(zawartosc) if zawartosc else []
        except ET.ParseError as exc:
            blad_xml = f"Nie udało się odczytać XML: {exc}"

    korekty = []
    if obj.invoice_hash:
        korekty = list(Faktura.objects.filter(hash_of_corrected_invoice=obj.invoice_hash))

    context = {
        "object": obj,
        "module_name": "Faktura",
        "list_url_name": "faktura_list",
        "plik_xml": plik_xml,
        "pozycje": pozycje,
        "blad_xml": blad_xml,
        "korekty": korekty,
    }
    return render(request, "kadry/faktura_detail.html", context)


def zalacznik_create(request, content_type_id, object_id):
    """Uniwersalne dodawanie pliku do dowolnego obiektu."""
    ct = get_object_or_404(ContentType, pk=content_type_id)
    obj = ct.get_object_for_this_type(pk=object_id)

    if request.method == "POST":
        form = ZalacznikForm(request.POST, request.FILES)
        if form.is_valid():
            zalacznik = form.save(commit=False)
            zalacznik.content_type = ct
            zalacznik.object_id = obj.pk
            zalacznik.save()
            django_messages.success(request, "Poprawnie przesłano załącznik.")
        else:
            django_messages.error(request, "Błąd dodawania załącznika.")

    if ct.model == "pracownik":
        return redirect("pracownik_detail", pk=obj.pk)
    elif ct.model == "pojazd":
        return redirect("pojazd_detail", pk=obj.pk)
    return redirect("dashboard")


def zalacznik_delete(request, pk):
    """Usuwanie załącznika wraz z fizycznym plikiem."""
    zalacznik = get_object_or_404(Zalacznik, pk=pk)
    ct_model = zalacznik.content_type.model
    obj_pk = zalacznik.object_id

    if request.method == "POST":
        zalacznik.plik.delete(save=False) # fizyczne usuwanie z dysku
        zalacznik.delete()
        django_messages.success(request, "Usunięto załącznik i plik z serwera.")

    if ct_model == "pracownik":
        return redirect("pracownik_detail", pk=obj_pk)
    elif ct_model == "pojazd":
        return redirect("pojazd_detail", pk=obj_pk)
    return redirect("dashboard")
